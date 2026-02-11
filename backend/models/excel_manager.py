"""
Excel Manager - Gestor centralizado de operaciones con Excel
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os
import shutil
from pathlib import Path

class ExcelManager:
    """Gestor centralizado para todas las operaciones con Excel"""
    
    def __init__(self, excel_path, backup_dir='backups'):
        self.excel_path = Path(excel_path)
        self.backup_dir = Path(backup_dir)
        self.backup_dir.mkdir(exist_ok=True)
        
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel no encontrado: {self.excel_path}")
    
    # ==================== BACKUPS ====================
    
    def create_backup(self):
        """Crear backup del Excel antes de modificar"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f'backup_{timestamp}.xlsx'
        backup_path = self.backup_dir / backup_name
        
        shutil.copy2(self.excel_path, backup_path)
        return str(backup_path)
    
    def cleanup_old_backups(self, days=30):
        """Eliminar backups antiguos"""
        cutoff = datetime.now() - timedelta(days=days)
        
        for backup in self.backup_dir.glob('backup_*.xlsx'):
            if datetime.fromtimestamp(backup.stat().st_mtime) < cutoff:
                backup.unlink()
    
    # ==================== PEDIDOS ====================
    
    def get_all_pedidos(self):
        """Obtener todos los pedidos"""
        wb = load_workbook(self.excel_path, data_only=True)
        ws = wb['2. Pedidos']
        
        pedidos = []
        fila = 4  # Primera fila de datos
        
        while ws[f'A{fila}'].value:
            pedido = {
                'id': ws[f'A{fila}'].value,
                'fecha_pago': self._format_date(ws[f'B{fila}'].value),
                'fecha_compromiso': self._format_date(ws[f'C{fila}'].value),
                'fecha_entrega_real': self._format_date(ws[f'D{fila}'].value),
                'dias_produccion': ws[f'E{fila}'].value,
                'dias_retraso': ws[f'F{fila}'].value,
                'sku': ws[f'G{fila}'].value,
                'producto': ws[f'H{fila}'].value,
                'talla': ws[f'I{fila}'].value,
                'personalizacion': ws[f'J{fila}'].value,
                'costo_person': ws[f'K{fila}'].value,
                'cliente': ws[f'L{fila}'].value,
                'telefono': ws[f'M{fila}'].value,
                'email': ws[f'N{fila}'].value,
                'direccion': ws[f'O{fila}'].value,
                'canal': ws[f'P{fila}'].value,
                'precio_producto': ws[f'Q{fila}'].value,
                'precio_person': ws[f'R{fila}'].value,
                'precio_envio': ws[f'S{fila}'].value,
                'precio_total': ws[f'T{fila}'].value,
                'costo_total': ws[f'U{fila}'].value,
                'ganancia': ws[f'V{fila}'].value,
                'banco': ws[f'W{fila}'].value,
                'estatus_pago': ws[f'X{fila}'].value,
                'estatus_produccion': ws[f'Y{fila}'].value
            }
            pedidos.append(pedido)
            fila += 1
        
        wb.close()
        return pedidos
    
    def get_pedido_by_id(self, pedido_id):
        """Obtener un pedido específico por ID"""
        pedidos = self.get_all_pedidos()
        for pedido in pedidos:
            if pedido['id'] == pedido_id:
                return pedido
        return None
    
    def create_pedido(self, data):
        """Crear un nuevo pedido"""
        self.create_backup()
        
        wb = load_workbook(self.excel_path)
        ws = wb['2. Pedidos']
        
        # Encontrar siguiente ID
        nuevo_id = self._get_next_id(ws)
        
        # Encontrar primera fila vacía
        fila = 4
        while ws[f'A{fila}'].value:
            fila += 1
        
        # Calcular fechas
        fecha_pago = datetime.now()
        dias_produccion = int(data.get('tiempo_estimado', '7').split()[0])
        fecha_compromiso = fecha_pago + timedelta(days=dias_produccion)
        
        # Obtener precios del catálogo
        ws_catalogo = wb['1. Catálogo']
        producto_base = self._get_precio_producto(ws_catalogo, data['producto_sku'])
        personalizacion_precio = self._get_precio_personalizacion(ws_catalogo, data.get('personalizacion_tipo'))
        
        precio_envio = int(data.get('costo_envio', 200))
        
        # Escribir datos
        self._write_pedido_row(ws, fila, {
            'id': nuevo_id,
            'fecha_pago': fecha_pago.strftime('%d/%m/%Y'),
            'fecha_compromiso': fecha_compromiso.strftime('%d/%m/%Y'),
            'sku': data['producto_sku'],
            'producto': data['producto_nombre'],
            'talla': data['talla'],
            'personalizacion': data.get('personalizacion_detalles', ''),
            'costo_person': personalizacion_precio,
            'cliente': data['nombre_cliente'],
            'telefono': data['telefono'],
            'email': data.get('email', ''),
            'direccion': data['direccion'],
            'canal': data['canal'],
            'precio_producto': producto_base,
            'precio_person': personalizacion_precio,
            'precio_envio': precio_envio,
            'banco': data['banco'],
            'estatus_pago': data['estatus_pago'],
            'estatus_produccion': 'En Producción'
        })
        
        wb.save(self.excel_path)
        wb.close()
        
        return {
            'id': nuevo_id,
            'fecha_entrega': fecha_compromiso.strftime('%d/%m/%Y'),
            'total': producto_base + personalizacion_precio + precio_envio
        }
    
    def update_pedido(self, pedido_id, data):
        """Actualizar un pedido existente"""
        self.create_backup()
        
        wb = load_workbook(self.excel_path)
        ws = wb['2. Pedidos']
        
        # Buscar pedido
        fila = 4
        encontrado = False
        
        while ws[f'A{fila}'].value:
            if ws[f'A{fila}'].value == pedido_id:
                encontrado = True
                
                # Actualizar campos permitidos
                if 'estatus_produccion' in data:
                    ws[f'Y{fila}'].value = data['estatus_produccion']
                
                if 'estatus_pago' in data:
                    ws[f'X{fila}'].value = data['estatus_pago']
                
                if 'direccion' in data:
                    ws[f'O{fila}'].value = data['direccion']
                
                if 'fecha_entrega_real' in data and data['fecha_entrega_real']:
                    ws[f'D{fila}'].value = data['fecha_entrega_real']
                    ws[f'D{fila}'].number_format = 'DD/MM/YYYY'
                
                wb.save(self.excel_path)
                wb.close()
                return True
            
            fila += 1
        
        wb.close()
        return False if not encontrado else None
    
    def delete_pedido(self, pedido_id):
        """Eliminar un pedido"""
        self.create_backup()
        
        wb = load_workbook(self.excel_path)
        ws = wb['2. Pedidos']
        
        fila = 4
        while ws[f'A{fila}'].value:
            if ws[f'A{fila}'].value == pedido_id:
                ws.delete_rows(fila)
                wb.save(self.excel_path)
                wb.close()
                return True
            fila += 1
        
        wb.close()
        return False
    
    def get_pedidos_pendientes(self):
        """Obtener pedidos que requieren atención"""
        pedidos = self.get_all_pedidos()
        pendientes = []
        
        for pedido in pedidos:
            es_pendiente = (
                'Bloqueado' in str(pedido.get('estatus_produccion', '')) or
                pedido.get('direccion') == 'Pendiente' or
                (isinstance(pedido.get('dias_retraso'), (int, float)) and pedido['dias_retraso'] > 0)
            )
            
            if es_pendiente:
                motivos = []
                if 'Bloqueado' in str(pedido.get('estatus_produccion', '')):
                    motivos.append('Sin dirección')
                if isinstance(pedido.get('dias_retraso'), (int, float)) and pedido['dias_retraso'] > 0:
                    motivos.append(f'{int(pedido["dias_retraso"])} días de retraso')
                
                pedido['motivo'] = ', '.join(motivos)
                pendientes.append(pedido)
        
        return pendientes
    
    # ==================== PRODUCTOS ====================
    
    def get_all_productos(self):
        """Obtener catálogo completo de productos"""
        wb = load_workbook(self.excel_path, data_only=True)
        ws = wb['1. Catálogo']
        
        productos = []
        for row in range(4, 20):
            if not ws[f'A{row}'].value:
                break
            
            productos.append({
                'sku': ws[f'A{row}'].value,
                'nombre': ws[f'B{row}'].value,
                'categoria': ws[f'C{row}'].value,
                'precio': ws[f'D{row}'].value,
                'costo_material': ws[f'E{row}'].value,
                'costo_mano_obra': ws[f'F{row}'].value,
                'costo_total': ws[f'G{row}'].value,
                'margen': ws[f'H{row}'].value,
                'margen_pct': ws[f'I{row}'].value,
                'tiempo': ws[f'J{row}'].value
            })
        
        wb.close()
        return productos
    
    # ==================== CLIENTES ====================
    
    def get_all_clientes(self):
        """Obtener base de datos de clientes con estadísticas"""
        from collections import defaultdict
        
        pedidos = self.get_all_pedidos()
        clientes_dict = defaultdict(lambda: {
            'nombre': '',
            'telefono': '',
            'email': '',
            'pedidos': 0,
            'total_gastado': 0,
            'ultimo_pedido': None
        })
        
        for pedido in pedidos:
            nombre = pedido.get('cliente')
            if not nombre:
                continue
            
            cliente = clientes_dict[nombre]
            cliente['nombre'] = nombre
            cliente['telefono'] = pedido.get('telefono') or cliente['telefono']
            cliente['email'] = pedido.get('email') or cliente['email']
            cliente['pedidos'] += 1
            
            total = pedido.get('precio_total')
            if isinstance(total, (int, float)):
                cliente['total_gastado'] += total
            
            fecha = pedido.get('fecha_pago')
            if not cliente['ultimo_pedido'] or (fecha and fecha > cliente['ultimo_pedido']):
                cliente['ultimo_pedido'] = fecha
        
        # Clasificar clientes
        clientes = []
        for cliente in clientes_dict.values():
            if cliente['pedidos'] >= 5:
                cliente['tipo'] = 'VIP'
            elif cliente['pedidos'] >= 3:
                cliente['tipo'] = 'Frecuente'
            else:
                cliente['tipo'] = 'Nuevo'
            
            clientes.append(cliente)
        
        return clientes
    
    # ==================== ESTADÍSTICAS ====================
    
    def get_estadisticas(self):
        """Obtener estadísticas generales del negocio"""
        pedidos = self.get_all_pedidos()
        
        total_pedidos = len(pedidos)
        ventas_totales = sum(p.get('precio_total', 0) for p in pedidos if isinstance(p.get('precio_total'), (int, float)))
        ganancia_neta = sum(p.get('ganancia', 0) for p in pedidos if isinstance(p.get('ganancia'), (int, float)))
        
        pedidos_pendientes = len([
            p for p in pedidos 
            if p.get('estatus_produccion') and ('Bloqueado' in str(p['estatus_produccion']) or 'Pendiente' in str(p['estatus_produccion']))
        ])
        
        margen_promedio = (ganancia_neta / ventas_totales * 100) if ventas_totales > 0 else 0
        
        return {
            'total_pedidos': total_pedidos,
            'ventas_totales': ventas_totales,
            'ganancia_neta': ganancia_neta,
            'pedidos_pendientes': pedidos_pendientes,
            'margen_promedio': margen_promedio
        }
    
    # ==================== UTILIDADES PRIVADAS ====================
    
    def _get_next_id(self, ws):
        """Obtener siguiente ID de pedido"""
        ultimo_id = 0
        fila = 4
        
        while ws[f'A{fila}'].value:
            id_actual = ws[f'A{fila}'].value
            if id_actual and isinstance(id_actual, str) and id_actual.startswith('P'):
                try:
                    numero = int(id_actual[1:])
                    ultimo_id = max(ultimo_id, numero)
                except:
                    pass
            fila += 1
        
        return f'P{str(ultimo_id + 1).zfill(4)}'
    
    def _get_precio_producto(self, ws_catalogo, sku):
        """Obtener precio de producto del catálogo"""
        for row in range(4, 20):
            if ws_catalogo[f'A{row}'].value == sku:
                return ws_catalogo[f'D{row}'].value or 0
        return 0
    
    def _get_precio_personalizacion(self, ws_catalogo, codigo):
        """Obtener precio de personalización del catálogo"""
        if not codigo or codigo == 'ninguna':
            return 0
        
        for row in range(14, 20):
            if ws_catalogo[f'A{row}'].value == codigo:
                return ws_catalogo[f'C{row}'].value or 0
        return 0
    
    def _write_pedido_row(self, ws, fila, data):
        """Escribir fila de pedido en Excel"""
        # Escribir valores
        ws[f'A{fila}'] = data['id']
        ws[f'B{fila}'] = data['fecha_pago']
        ws[f'C{fila}'] = data['fecha_compromiso']
        ws[f'D{fila}'] = ''
        ws[f'E{fila}'] = f'=C{fila}-B{fila}'
        ws[f'F{fila}'] = f'=IF(D{fila}="",IF(TODAY()>C{fila},TODAY()-C{fila},0),IF(D{fila}>C{fila},D{fila}-C{fila},0))'
        ws[f'G{fila}'] = data['sku']
        ws[f'H{fila}'] = data['producto']
        ws[f'I{fila}'] = data['talla']
        ws[f'J{fila}'] = data['personalizacion']
        ws[f'K{fila}'] = data['costo_person']
        ws[f'L{fila}'] = data['cliente']
        ws[f'M{fila}'] = data['telefono']
        ws[f'N{fila}'] = data['email']
        ws[f'O{fila}'] = data['direccion']
        ws[f'P{fila}'] = data['canal']
        ws[f'Q{fila}'] = data['precio_producto']
        ws[f'R{fila}'] = data['precio_person']
        ws[f'S{fila}'] = data['precio_envio']
        ws[f'T{fila}'] = f'=Q{fila}+R{fila}+S{fila}'
        ws[f'U{fila}'] = f'=VLOOKUP(G{fila},\'1. Catálogo\'!$A$4:$G$9,7,FALSE)+K{fila}+S{fila}'
        ws[f'V{fila}'] = f'=T{fila}-U{fila}'
        ws[f'W{fila}'] = data['banco']
        ws[f'X{fila}'] = data['estatus_pago']
        ws[f'Y{fila}'] = data['estatus_produccion']
        
        # Aplicar formatos
        for col in ['B', 'C', 'D']:
            ws[f'{col}{fila}'].number_format = 'DD/MM/YYYY'
        
        for col in ['K', 'Q', 'R', 'S', 'T', 'U', 'V']:
            ws[f'{col}{fila}'].number_format = '$#,##0'
    
    def _format_date(self, value):
        """Formatear fecha para JSON"""
        if value is None or value == '':
            return None
        if isinstance(value, datetime):
            return value.strftime('%d/%m/%Y')
        if isinstance(value, str):
            return value
        return str(value)
